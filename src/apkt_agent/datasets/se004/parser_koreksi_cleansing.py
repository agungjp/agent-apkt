"""Parser for Koreksi Cleansing Excel files.

This parser handles Excel files from the Koreksi Cleansing report.
The structure may differ from SE004 reports.
"""

import re
from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import load_workbook

from ...logging_ import get_logger


# Unit code to name mapping for fallback extraction
UNIT_CODE_TO_NAME = {
    "WIL_ACEH": "WILAYAH ACEH",
    "WIL_SUMUT": "WILAYAH SUMATERA UTARA",
    "WIL_SUMBAR": "WILAYAH SUMATERA BARAT",
    "WIL_S2JB": "WILAYAH SUMATERA SELATAN, JAMBI & BENGKULU (S2JB)",
    "WIL_BABEL": "WILAYAH BANGKA BELITUNG",
    "DIST_LAMPUNG": "DISTRIBUSI LAMPUNG",
    "WIL_RIAUKEPRI": "WILAYAH RIAU DAN KEPULAUAN RIAU",
    "WIL_KALBAR": "WILAYAH KALIMANTAN BARAT",
    "WIL_KALSELTENG": "WILAYAH KALIMANTAN SELATAN DAN TENGAH",
    "WIL_KALTIM": "WILAYAH KALIMANTAN TIMUR",
    "REG_SUMKAL": "REGIONAL SUMKAL",
}


def list_excel_files(raw_excel_dir: Path) -> List[Path]:
    """List all Excel files in directory, sorted by name.
    
    Args:
        raw_excel_dir: Path to directory containing Excel files
        
    Returns:
        List of Path objects for each .xlsx file, sorted alphabetically
    """
    if not raw_excel_dir.exists():
        return []
    
    xlsx_files = list(raw_excel_dir.glob("*.xlsx"))
    # Filter out temp files (starting with ~$)
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]
    return sorted(xlsx_files, key=lambda p: p.name)


def _extract_unit_from_filename(filename: str) -> Optional[str]:
    """Extract unit name from filename as fallback.
    
    Args:
        filename: Like "koreksi_cleansing_202508_WIL_ACEH.xlsx"
        
    Returns:
        Unit name like "WILAYAH ACEH" or None
    """
    # Remove extension
    name = filename.replace(".xlsx", "").replace(".xls", "")
    
    # Try to find unit code in filename
    for code, full_name in UNIT_CODE_TO_NAME.items():
        if code in name:
            return full_name
    
    # Fallback: try to extract last part after underscore
    parts = name.split("_")
    if len(parts) >= 2:
        # Get last 2 parts, e.g., WIL_ACEH
        potential_code = "_".join(parts[-2:])
        if potential_code in UNIT_CODE_TO_NAME:
            return UNIT_CODE_TO_NAME[potential_code]
        # Try last part only
        if parts[-1] in UNIT_CODE_TO_NAME:
            return UNIT_CODE_TO_NAME[parts[-1]]
    
    return None


def _extract_period_from_filename(filename: str) -> Optional[str]:
    """Extract period (YYYYMM) from filename.
    
    Args:
        filename: Like "koreksi_cleansing_202508_WIL_ACEH.xlsx"
        
    Returns:
        Period like "202508" or None
    """
    # Look for 6 digit pattern YYYYMM
    match = re.search(r'_(\d{6})_', filename)
    if match:
        return match.group(1)
    return None


def _find_header_row(ws) -> Optional[int]:
    """Find the header row in the worksheet.
    
    Args:
        ws: openpyxl worksheet
        
    Returns:
        Row number (1-indexed) of header row, or None if not found
    """
    # Look for common header patterns
    header_patterns = [
        "No",
        "NO",
        "Nomor",
        "Unit",
        "Tanggal",
        "Kode Gangguan",
    ]
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20), start=1):
        for cell in row:
            if cell.value:
                cell_str = str(cell.value).strip()
                for pattern in header_patterns:
                    if pattern.lower() == cell_str.lower():
                        return row_idx
    
    return None


def parse_excel_file(filepath: Path) -> pd.DataFrame:
    """Parse a single Koreksi Cleansing Excel file to DataFrame.
    
    This function uses a flexible approach:
    1. First tries to find header row
    2. If not found, reads from row 1
    3. Adds metadata columns (unit_induk, period_ym, source_file)
    
    Args:
        filepath: Path to Excel file
        
    Returns:
        DataFrame with parsed data
    """
    logger = get_logger()
    logger.info(f"Parsing: {filepath.name}")
    
    # Extract metadata from filename
    unit_induk = _extract_unit_from_filename(filepath.name)
    period_ym = _extract_period_from_filename(filepath.name)
    
    try:
        # Load workbook to find header row
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # Find header row
        header_row = _find_header_row(ws)
        if header_row is None:
            header_row = 1  # Default to first row
            logger.warning(f"Header row not found, using row 1")
        
        wb.close()
        
        # Read with pandas, using found header row
        df = pd.read_excel(
            filepath,
            header=header_row - 1,  # pandas uses 0-indexed
            engine='openpyxl'
        )
        
        # Remove empty rows
        df = df.dropna(how='all')
        
        # Add metadata columns
        df['unit_induk'] = unit_induk or "UNKNOWN"
        df['period_ym'] = period_ym or "UNKNOWN"
        df['source_file'] = filepath.name
        
        logger.info(f"Parsed {len(df)} rows from {filepath.name}")
        return df
        
    except Exception as e:
        logger.error(f"Failed to parse {filepath.name}: {e}")
        raise


def parse_all_excel_files(excel_dir: Path) -> pd.DataFrame:
    """Parse all Excel files in directory and combine into single DataFrame.
    
    Args:
        excel_dir: Path to directory containing Excel files
        
    Returns:
        Combined DataFrame with all data
    """
    logger = get_logger()
    
    excel_files = list_excel_files(excel_dir)
    logger.info(f"Found {len(excel_files)} Excel files to parse")
    
    if not excel_files:
        logger.warning("No Excel files found")
        return pd.DataFrame()
    
    all_dfs = []
    for filepath in excel_files:
        try:
            df = parse_excel_file(filepath)
            if len(df) > 0:
                all_dfs.append(df)
        except Exception as e:
            logger.error(f"Skipping {filepath.name}: {e}")
            continue
    
    if not all_dfs:
        logger.warning("No data parsed from any file")
        return pd.DataFrame()
    
    # Combine all DataFrames
    combined = pd.concat(all_dfs, ignore_index=True)
    logger.info(f"Combined DataFrame: {len(combined)} total rows from {len(all_dfs)} files")
    
    return combined


def save_csv_indonesian_format(df: pd.DataFrame, output_path: Path) -> Path:
    """Save DataFrame to CSV with Indonesian format (semicolon separator).
    
    Args:
        df: DataFrame to save
        output_path: Output file path
        
    Returns:
        Path to saved CSV file
    """
    logger = get_logger()
    
    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Save with semicolon separator (common in Indonesian Excel)
    df.to_csv(
        output_path,
        sep=';',
        index=False,
        encoding='utf-8-sig',  # UTF-8 with BOM for Excel compatibility
    )
    
    logger.info(f"Saved CSV with Indonesian format to: {output_path}")
    return output_path
