"""Parser for SE004 Kumulatif Excel files."""

import re
from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import load_workbook

from .schema import (
    SE004_KUMULATIF_COLUMNS,
    parse_indonesian_number,
    parse_period_label_to_ym,
    format_indonesian_number,
    parse_tanggal_to_ddmmyyyy,
)
from ...logging_ import get_logger


def list_excel_files(raw_excel_dir: Path) -> List[Path]:
    """List all Excel files in directory, sorted by name.
    
    Args:
        raw_excel_dir: Path to directory containing Excel files
        
    Returns:
        List of Path objects for each .xlsx file, sorted alphabetically
    """
    if not raw_excel_dir.exists():
        return []
    
    xlsx_files = list(raw_excel_dir.glob("*.xlsx"))
    # Filter out temp files (starting with ~$)
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]
    return sorted(xlsx_files, key=lambda p: p.name)


def _find_cell_by_text(ws, search_text: str, regex: bool = False) -> Optional[tuple]:
    """Find cell containing specific text.
    
    Args:
        ws: openpyxl worksheet
        search_text: Text to search for
        regex: If True, use regex matching
        
    Returns:
        Tuple of (row, col, cell_value) or None if not found
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if cell.value:
                cell_str = str(cell.value).strip()
                if regex:
                    if re.search(search_text, cell_str, re.IGNORECASE):
                        return (row_idx, col_idx, cell_str)
                else:
                    if search_text.lower() in cell_str.lower():
                        return (row_idx, col_idx, cell_str)
    return None


def _extract_after_colon(text: str) -> str:
    """Extract text after colon and strip it.
    
    Args:
        text: Text like "UNIT INDUK : WILAYAH ACEH"
        
    Returns:
        Text after colon, e.g., "WILAYAH ACEH"
    """
    if not text or ":" not in text:
        return text
    return text.split(":", 1)[1].strip()


# Unit code to name mapping for fallback extraction
UNIT_CODE_TO_NAME = {
    "WIL_ACEH": "WILAYAH ACEH",
    "WIL_SUMUT": "WILAYAH SUMATERA UTARA",
    "WIL_SUMBAR": "WILAYAH SUMATERA BARAT",
    "WIL_S2JB": "WILAYAH SUMATERA SELATAN, JAMBI & BENGKULU (S2JB)",
    "WIL_BABEL": "WILAYAH BANGKA BELITUNG",
    "DIST_LAMPUNG": "DISTRIBUSI LAMPUNG",
    "WIL_RIAUKEPRI": "WILAYAH RIAU DAN KEPULAUAN RIAU",
    "WIL_KALBAR": "WILAYAH KALIMANTAN BARAT",
    "WIL_KALSELTENG": "WILAYAH KALIMANTAN SELATAN DAN TENGAH",
    "WIL_KALTIM": "WILAYAH KALIMANTAN TIMUR",
    "REG_SUMKAL": "REGIONAL SUMKAL",
}


def _extract_unit_from_filename(filename: str) -> Optional[str]:
    """Extract unit name from filename as fallback.
    
    Args:
        filename: Like "se004_bulanan_202508_WIL_ACEH.xlsx"
        
    Returns:
        Unit name like "WILAYAH ACEH" or None
    """
    # Remove extension
    name = filename.replace(".xlsx", "").replace(".xls", "")
    
    # Try to find unit code in filename
    for code, full_name in UNIT_CODE_TO_NAME.items():
        if code in name:
            return full_name
    
    # Fallback: try to extract last part after underscore
    parts = name.split("_")
    if len(parts) >= 2:
        # Get last 2 parts, e.g., WIL_ACEH
        potential_code = "_".join(parts[-2:])
        if potential_code in UNIT_CODE_TO_NAME:
            return UNIT_CODE_TO_NAME[potential_code]
        # Try last part only
        if parts[-1] in UNIT_CODE_TO_NAME:
            return UNIT_CODE_TO_NAME[parts[-1]]
    
    return None

def _get_cell_value(ws, row: int, col: int) -> Any:
    """Get cell value safely.
    
    Args:
        ws: openpyxl worksheet
        row: Row number (1-indexed)
        col: Column number (1-indexed)
        
    Returns:
        Cell value or None
    """
    try:
        return ws.cell(row=row, column=col).value
    except Exception:
        return None


def _parse_numeric_value(value: Any) -> Optional[float]:
    """Parse a cell value to numeric, handling Indonesian format.
    
    Args:
        value: Cell value (can be number or string)
        
    Returns:
        Float value or None
    """
    if value is None:
        return None
    
    if isinstance(value, (int, float)):
        return float(value)
    
    if isinstance(value, str):
        value = value.strip()
        if value == "-" or value == "":
            return None
        return parse_indonesian_number(value)
    
    return None


def _determine_row_type(kode: Any, penyebab: str, numeric_values: List[Any]) -> str:
    """Determine the type of row.
    
    Args:
        kode: Kode value
        penyebab: Penyebab gangguan text
        numeric_values: List of numeric column values
        
    Returns:
        Row type: "detail", "subtotal", "total", or "group"
    """
    penyebab_upper = (penyebab or "").upper().strip()
    
    # Check for total rows
    if kode is None or str(kode).strip() == "":
        if "TOTAL" in penyebab_upper:
            return "total"
        return "subtotal"
    
    # Check if all numeric values are empty/None
    all_empty = all(v is None for v in numeric_values)
    if all_empty:
        return "group"
    
    return "detail"


def parse_se004_kumulatif_xlsx(file_path: Path) -> pd.DataFrame:
    """Parse a single SE004 Kumulatif Excel file.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        pandas DataFrame with parsed data
    """
    logger = get_logger()
    logger.info(f"Parsing: {file_path.name}")
    
    # Load workbook
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # === EXTRACT METADATA ===
    metadata = {}
    
    # 1. unit_induk: Find cell containing "UNIT INDUK"
    unit_cell = _find_cell_by_text(ws, "UNIT INDUK")
    if unit_cell:
        extracted_unit = _extract_after_colon(unit_cell[2])
        # If empty after colon (merged cell issue), try to extract from filename
        if not extracted_unit or extracted_unit.strip() == "":
            # Filename like: se004_bulanan_202508_WIL_ACEH.xlsx -> WIL_ACEH -> WILAYAH ACEH
            extracted_unit = _extract_unit_from_filename(file_path.name)
        metadata["unit_induk"] = extracted_unit
    else:
        # Fallback to filename
        metadata["unit_induk"] = _extract_unit_from_filename(file_path.name)
    
    # 2. period_label: Find Indonesian month + year pattern
    period_cell = _find_cell_by_text(ws, r"(Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Desember)\s+\d{4}", regex=True)
    if period_cell:
        # Extract just the month year part
        match = re.search(r"(Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Desember)\s+\d{4}", period_cell[2], re.IGNORECASE)
        metadata["period_label"] = match.group(0) if match else period_cell[2]
    else:
        metadata["period_label"] = None
    
    # 3. period_ym: Convert from period_label
    metadata["period_ym"] = parse_period_label_to_ym(metadata["period_label"])
    
    # 4. jumlah_pelanggan: Find "Jumlah Pelanggan" then get value from next cell (C2)
    jml_plg_cell = _find_cell_by_text(ws, "Jumlah Pelanggan")
    if jml_plg_cell:
        row, col, text = jml_plg_cell
        # Value is in the NEXT column (C2)
        next_val = _get_cell_value(ws, row, col + 1)
        if next_val:
            # Parse as string to handle Indonesian format
            val_str = str(next_val).strip()
            metadata["jumlah_pelanggan"] = parse_indonesian_number(val_str)
        else:
            # Fallback: check if value is after colon
            if ":" in text:
                val_text = _extract_after_colon(text)
                metadata["jumlah_pelanggan"] = parse_indonesian_number(val_text)
            else:
                metadata["jumlah_pelanggan"] = None
    else:
        metadata["jumlah_pelanggan"] = None
    
    # 5. saidi_total: Find "SAIDI :" pattern (with colon) - value in C2 (Jam/Plg), C4 (Menit/Plg)
    # Use more specific search to avoid matching "LAPORAN SAIDI SAIFI"
    saidi_cell = _find_cell_by_text(ws, "SAIDI :")
    if not saidi_cell:
        saidi_cell = _find_cell_by_text(ws, "SAIDI:")
    if saidi_cell:
        row, col, text = saidi_cell
        # Jam/Plg is in next column (usually C2)
        jam_val = _get_cell_value(ws, row, col + 1)
        if jam_val:
            val_str = str(jam_val).strip()
            metadata["saidi_total"] = parse_indonesian_number(val_str)
        else:
            metadata["saidi_total"] = None
        # Menit/Plg is usually in C4 (col + 3)
        menit_val = _get_cell_value(ws, row, col + 3)
        if menit_val:
            val_str = str(menit_val).strip()
            metadata["saidi_total_menit"] = parse_indonesian_number(val_str)
        else:
            metadata["saidi_total_menit"] = None
    else:
        metadata["saidi_total"] = None
        metadata["saidi_total_menit"] = None
    
    # 6. saifi_total: Find "SAIFI :" pattern (with colon) - value in C2 (Kali/Plg)
    saifi_cell = _find_cell_by_text(ws, "SAIFI :")
    if not saifi_cell:
        saifi_cell = _find_cell_by_text(ws, "SAIFI:")
    if saifi_cell:
        row, col, text = saifi_cell
        # Value is in next column (C2)
        next_val = _get_cell_value(ws, row, col + 1)
        if next_val:
            val_str = str(next_val).strip()
            metadata["saifi_total"] = parse_indonesian_number(val_str)
        else:
            metadata["saifi_total"] = None
    else:
        metadata["saifi_total"] = None
    
    # 7. tanggal_cetak: Find "Tanggal Penarikan" or "Tanggal Cetak", format to dd/mm/yyyy
    tgl_cell = _find_cell_by_text(ws, "Tanggal Penarikan")
    if not tgl_cell:
        tgl_cell = _find_cell_by_text(ws, "Tanggal Cetak")
    if tgl_cell:
        raw_date = _extract_after_colon(tgl_cell[2])
        metadata["tanggal_cetak"] = parse_tanggal_to_ddmmyyyy(raw_date)
    else:
        metadata["tanggal_cetak"] = None
    
    # === FIND TABLE HEADER ===
    header_cell = _find_cell_by_text(ws, "NO. KODE")
    if not header_cell:
        header_cell = _find_cell_by_text(ws, "NO.KODE")
    if not header_cell:
        header_cell = _find_cell_by_text(ws, "KODE")
    
    if not header_cell:
        wb.close()
        # Return empty DataFrame with metadata if no table found
        return pd.DataFrame([{
            **metadata,
            "kode": None,
            "penyebab_gangguan": None,
            "jml_plg_padam": None,
            "jam_x_jml_plg_padam": None,
            "saidi_jam": None,
            "saifi_kali": None,
            "jumlah_gangguan_kali": None,
            "lama_padam_jam": None,
            "kwh_tak_tersalurkan": None,
            "row_type": "metadata_only",
            "source_file": file_path.name,
        }])
    
    header_row = header_cell[0]
    
    # === BUILD COLUMN MAPPING ===
    # Expected headers (flexible matching)
    header_patterns = {
        "kode": [r"NO\.?\s*KODE", r"KODE"],
        "penyebab_gangguan": [r"PENYEBAB\s*GANGGUAN", r"PENYEBAB"],
        "jml_plg_padam": [r"JML\.?\s*PLG\.?\s*PADAM", r"PELANGGAN\s*PADAM"],
        "jam_x_jml_plg_padam": [r"JAM\s*X\s*JML", r"JAM.*PLG.*PADAM"],
        "saidi_jam": [r"SAIDI.*JAM", r"SAIDI"],
        "saifi_kali": [r"SAIFI.*KALI", r"SAIFI"],
        "jumlah_gangguan_kali": [r"JML\.?\s*GANGGUAN", r"GANGGUAN.*KALI"],
        "lama_padam_jam": [r"LAMA\s*PADAM", r"PADAM.*JAM"],
        "kwh_tak_tersalurkan": [r"KWH\s*TAK", r"KWH"],
    }
    
    col_mapping = {}
    for col_idx in range(1, 20):  # Check first 20 columns
        cell_val = _get_cell_value(ws, header_row, col_idx)
        if cell_val:
            cell_str = str(cell_val).upper().strip()
            for field, patterns in header_patterns.items():
                if field not in col_mapping:
                    for pattern in patterns:
                        if re.search(pattern, cell_str, re.IGNORECASE):
                            col_mapping[field] = col_idx
                            break
    
    # === PARSE DATA ROWS ===
    rows = []
    data_start_row = header_row + 1
    
    # Find end of table (look for SAIDI: or multiple empty rows)
    max_rows = 500  # Safety limit
    empty_row_count = 0
    
    for row_idx in range(data_start_row, data_start_row + max_rows):
        # Check first cell for end markers
        first_cell = _get_cell_value(ws, row_idx, 1)
        first_str = str(first_cell).strip().upper() if first_cell else ""
        
        # End markers
        if first_str.startswith("SAIDI:") or first_str.startswith("SAIFI:"):
            break
        
        # Check if row is empty
        kode_col = col_mapping.get("kode", 1)
        penyebab_col = col_mapping.get("penyebab_gangguan", 2)
        
        kode_val = _get_cell_value(ws, row_idx, kode_col)
        penyebab_val = _get_cell_value(ws, row_idx, penyebab_col)
        
        if not kode_val and not penyebab_val:
            empty_row_count += 1
            if empty_row_count >= 3:
                break
            continue
        else:
            empty_row_count = 0
        
        # Extract row data
        row_data = {**metadata}
        
        # Kode
        row_data["kode"] = kode_val
        
        # Penyebab gangguan
        row_data["penyebab_gangguan"] = str(penyebab_val).strip() if penyebab_val else None
        
        # Numeric columns
        numeric_fields = [
            "jml_plg_padam",
            "jam_x_jml_plg_padam", 
            "saidi_jam",
            "saifi_kali",
            "jumlah_gangguan_kali",
            "lama_padam_jam",
            "kwh_tak_tersalurkan",
        ]
        
        numeric_values = []
        for field in numeric_fields:
            col = col_mapping.get(field)
            if col:
                val = _get_cell_value(ws, row_idx, col)
                parsed = _parse_numeric_value(val)
                row_data[field] = parsed
                numeric_values.append(parsed)
            else:
                row_data[field] = None
                numeric_values.append(None)
        
        # Determine row type
        row_data["row_type"] = _determine_row_type(
            row_data["kode"],
            row_data["penyebab_gangguan"],
            numeric_values
        )
        
        # Source file
        row_data["source_file"] = file_path.name
        
        rows.append(row_data)
    
    wb.close()
    
    if not rows:
        # Return at least one row with metadata
        return pd.DataFrame([{
            **metadata,
            "kode": None,
            "penyebab_gangguan": None,
            "jml_plg_padam": None,
            "jam_x_jml_plg_padam": None,
            "saidi_jam": None,
            "saifi_kali": None,
            "jumlah_gangguan_kali": None,
            "lama_padam_jam": None,
            "kwh_tak_tersalurkan": None,
            "row_type": "no_data",
            "source_file": file_path.name,
        }])
    
    # Create DataFrame with proper column order
    df = pd.DataFrame(rows)
    
    # Ensure all columns exist
    for col in SE004_KUMULATIF_COLUMNS:
        if col not in df.columns:
            df[col] = None
    
    # Reorder columns
    df = df[SE004_KUMULATIF_COLUMNS]
    
    logger.info(f"Parsed {len(df)} rows from {file_path.name}")
    
    return df


def parse_all_excel_files(raw_excel_dir: Path) -> pd.DataFrame:
    """Parse all Excel files in directory and combine into one DataFrame.
    
    Args:
        raw_excel_dir: Path to directory containing Excel files
        
    Returns:
        Combined pandas DataFrame
    """
    logger = get_logger()
    files = list_excel_files(raw_excel_dir)
    
    logger.info(f"Found {len(files)} Excel files to parse")
    
    if not files:
        return pd.DataFrame(columns=SE004_KUMULATIF_COLUMNS)
    
    dfs = []
    for file_path in files:
        try:
            df = parse_se004_kumulatif_xlsx(file_path)
            dfs.append(df)
        except Exception as e:
            logger.error(f"Error parsing {file_path.name}: {e}")
            continue
    
    if not dfs:
        return pd.DataFrame(columns=SE004_KUMULATIF_COLUMNS)
    
    # Concatenate all DataFrames
    combined_df = pd.concat(dfs, ignore_index=True)
    
    logger.info(f"Combined DataFrame: {len(combined_df)} total rows from {len(dfs)} files")
    
    return combined_df


def save_csv_indonesian_format(df: pd.DataFrame, output_path: Path) -> Path:
    """Save DataFrame to CSV with Indonesian number format.
    
    Uses semicolon (;) as delimiter because comma is used as decimal separator.
    Numbers are formatted with dot (.) for thousands, comma (,) for decimals.
    
    Args:
        df: DataFrame to save
        output_path: Path to save CSV file
        
    Returns:
        Path to saved CSV file
    """
    logger = get_logger()
    
    # Define numeric columns that need formatting
    numeric_columns = [
        "jumlah_pelanggan",
        "saidi_total",
        "saidi_total_menit",
        "saifi_total",
        "jml_plg_padam",
        "jam_x_jml_plg_padam",
        "saidi_jam",
        "saifi_kali",
        "jumlah_gangguan_kali",
        "lama_padam_jam",
        "kwh_tak_tersalurkan",
    ]
    
    # Create a copy to avoid modifying original
    df_out = df.copy()
    
    # Strip and clean tanggal_cetak column
    if "tanggal_cetak" in df_out.columns:
        df_out["tanggal_cetak"] = df_out["tanggal_cetak"].apply(
            lambda x: str(x).strip() if pd.notna(x) and x else ""
        )
    
    # Format numeric columns to Indonesian format
    for col in numeric_columns:
        if col in df_out.columns:
            df_out[col] = df_out[col].apply(
                lambda x: format_indonesian_number(x, decimals=4) if pd.notna(x) else ""
            )
    
    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Save with semicolon delimiter (because comma is decimal separator)
    df_out.to_csv(output_path, index=False, sep=";", encoding="utf-8-sig")
    
    logger.info(f"Saved CSV with Indonesian format to: {output_path}")
    
    return output_path


# Legacy class for backward compatibility
class SE004Parser:
    """Parser for SE004 Excel files (legacy)."""
    
    def __init__(self):
        """Initialize parser."""
        self.logger = get_logger()
    
    def parse_excel(self, file_path: Path) -> List[Dict[str, Any]]:
        """Parse Excel file."""
        df = parse_se004_kumulatif_xlsx(file_path)
        return df.to_dict('records')
